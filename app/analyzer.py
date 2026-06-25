import openpyxl
from config import (
    CUSTOMER_COLS, REGION_COLS, FUEL_DECREASE_BUCKETS,
    FUEL_MAIN_THRESHOLD_GAL, NON_FUEL_ALERT_PCT,
    METRIC_SHEETS, FUEL_SHEET, METRIC_UNITS,
    CUSTOMER_DATA_START_ROW, REGION_DATA_START_ROW
)

NEWLY_DARK_MIN_GAL = 1000
FUEL_MAIN_MIN_CURRENT_GAL = 5000


def safe_float(val):
    if val is None or val == "" or val == "%" or val == "#":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def clean_field(val):
    if val is None or val == "#" or str(val).strip() == "#":
        return None
    return str(val).strip() if val else None


def pct_change(current, prior):
    if prior is None or prior == 0 or current is None:
        return None
    return ((current - prior) / prior) * 100


def get_report_date(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 2:
            if row[1]:
                return str(row[1])
    return "Unknown"


def get_fuel_bucket(pct):
    if pct is None:
        return None
    abs_pct = abs(pct)
    for low, high, label in FUEL_DECREASE_BUCKETS:
        if low <= abs_pct < high:
            return label
    return None


def find_last_known_value(row, start_col, end_col):
    for col in range(start_col, end_col):
        if col < len(row) and row[col] is not None:
            try:
                return float(row[col])
            except (TypeError, ValueError):
                continue
    return None


def analyze_customer_fuel(ws):
    main_alerts = []
    secondary_alerts = []
    newly_dark = []
    increases = []
    C = CUSTOMER_COLS

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < CUSTOMER_DATA_START_ROW:
            continue

        region = clean_field(row[C["region"]])
        customer = clean_field(row[C["customer"]])
        salesperson = clean_field(row[C["salesperson"]])
        area_mgr = clean_field(row[C["area_manager"]])
        avg_13wk = safe_float(row[C["avg_13wk"]])
        current_week = safe_float(row[C["current_week"]])
        prior_week = safe_float(row[C["prior_week"]])
        yoy_avg = safe_float(row[C["yoy_avg"]])

        if not customer or customer in ("Customer", ""):
            continue

        if current_week is None and avg_13wk is not None and avg_13wk >= NEWLY_DARK_MIN_GAL:
            last_known = find_last_known_value(row, C["current_week"], C["yoy_avg"])
            newly_dark.append({
                "customer": customer,
                "region": region,
                "salesperson": salesperson,
                "area_manager": area_mgr,
                "avg_13wk": avg_13wk,
                "last_known_vol": last_known,
            })
            continue

        if current_week is None or prior_week is None:
            continue

        wow_pct = pct_change(current_week, prior_week)
        if wow_pct is None:
            continue

        vol_change = current_week - prior_week
        bucket = get_fuel_bucket(wow_pct)
        is_main = avg_13wk is not None and avg_13wk >= FUEL_MAIN_THRESHOLD_GAL

        record = {
            "customer": customer,
            "region": region,
            "salesperson": salesperson,
            "area_manager": area_mgr,
            "avg_13wk": avg_13wk,
            "current_week": current_week,
            "prior_week": prior_week,
            "wow_pct": wow_pct,
            "vol_change": vol_change,
            "bucket": bucket,
            "yoy_avg": yoy_avg,
        }

        if wow_pct < 0 and bucket is not None:
            if is_main:
                if current_week >= FUEL_MAIN_MIN_CURRENT_GAL:
                    main_alerts.append(record)
            else:
                secondary_alerts.append(record)
        elif wow_pct > 0 and vol_change >= 5000:
            record["suppress_pct"] = wow_pct > 100
            increases.append(record)

    main_alerts.sort(key=lambda x: x["vol_change"])
    secondary_alerts.sort(key=lambda x: x["vol_change"])
    newly_dark.sort(key=lambda x: x["avg_13wk"] or 0, reverse=True)
    increases.sort(key=lambda x: x["vol_change"], reverse=True)

    increase_customers = {a["customer"] for a in increases}
    newly_dark = [a for a in newly_dark if a["customer"] not in increase_customers]
    return main_alerts, secondary_alerts, newly_dark, increases

def analyze_customer_non_fuel(ws, metric_name):
    alerts = []
    C = CUSTOMER_COLS
    unit = METRIC_UNITS.get(metric_name, "")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < CUSTOMER_DATA_START_ROW:
            continue

        customer = clean_field(row[C["customer"]])
        region = clean_field(row[C["region"]])
        salesperson = clean_field(row[C["salesperson"]])
        area_mgr = clean_field(row[C["area_manager"]])
        avg_13wk = safe_float(row[C["avg_13wk"]])
        current_week = safe_float(row[C["current_week"]])
        prior_week = safe_float(row[C["prior_week"]])

        if not customer or customer in ("Customer", ""):
            continue
        if current_week is None or prior_week is None:
            continue

        wow_pct = pct_change(current_week, prior_week)
        if wow_pct is None:
            continue
        if metric_name == "TCE Spend per Truck" and current_week is not None and current_week < 0:
            continue

        abs_change = abs(current_week - prior_week) if current_week is not None and prior_week is not None else 0
        min_current = {"Tires": 15, "PM": 5, "Labor Hrs": 10, "TCE Spend per Truck": 500}.get(metric_name, 0)
        min_change = {"Tires": 15, "PM": 5, "Labor Hrs": 10, "TCE Spend per Truck": 500}.get(metric_name, 0)
        if abs_change >= min_change and current_week >= min_current:
            alerts.append({
                "customer": customer,
                "region": region,
                "salesperson": salesperson,
                "area_manager": area_mgr,
                "metric": metric_name,
                "unit": unit,
                "avg_13wk": avg_13wk,
                "current_week": current_week,
                "prior_week": prior_week,
                "wow_pct": wow_pct,
                "direction": "decrease" if wow_pct < 0 else "increase",
            })

    alerts.sort(key=lambda x: x["wow_pct"])
    return alerts


def analyze_region_fuel(ws):
    regions = []
    C = REGION_COLS

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < REGION_DATA_START_ROW:
            continue

        label = clean_field(row[C["customer"]])
        salesperson = clean_field(row[C["salesperson"]])
        avg_13wk = safe_float(row[C["avg_13wk"]])
        curr_ov_avg = safe_float(row[C["curr_ov_avg"]])
        current_week = safe_float(row[C["current_week"]])
        prior_week = safe_float(row[C["prior_week"]])
        yoy_avg = safe_float(row[C["yoy_avg"]])

        if not label or label in ("Customer", ""):
            continue

        wow_pct = pct_change(current_week, prior_week)
        vol_change = (current_week - prior_week) if (current_week and prior_week) else None

        regions.append({
            "label": label,
            "salesperson": salesperson,
            "avg_13wk": avg_13wk,
            "curr_ov_avg": curr_ov_avg,
            "current_week": current_week,
            "prior_week": prior_week,
            "wow_pct": wow_pct,
            "vol_change": vol_change,
            "yoy_avg": yoy_avg,
        })

    return regions


def analyze_reports(customer_path, region_path):
    results = {
        "report_date": None,
        "region_summary": [],
        "fuel_main_alerts": [],
        "fuel_secondary": [],
        "fuel_newly_dark": [],
        "fuel_increases": [],
        "non_fuel_alerts": {},
        "errors": [],
    }

    try:
        if customer_path:
            wb_customer = openpyxl.load_workbook(customer_path, data_only=True)
            results["report_date"] = get_report_date(wb_customer[FUEL_SHEET])

            if FUEL_SHEET in wb_customer.sheetnames:
                ws = wb_customer[FUEL_SHEET]
                main, secondary, newly_dark, increases = analyze_customer_fuel(ws)
                results["fuel_main_alerts"] = main
                results["fuel_secondary"] = secondary
                results["fuel_newly_dark"] = newly_dark
                results["fuel_increases"] = increases

            for sheet in METRIC_SHEETS:
                if sheet == FUEL_SHEET:
                    continue
                if sheet in wb_customer.sheetnames:
                    ws = wb_customer[sheet]
                    alerts = analyze_customer_non_fuel(ws, sheet)
                    results["non_fuel_alerts"][sheet] = alerts

    except Exception as e:
        results["errors"].append(f"Customer report error: {str(e)}")

    try:
        if region_path:
            wb_region = openpyxl.load_workbook(region_path, data_only=True)

            if not results["report_date"]:
                results["report_date"] = get_report_date(wb_region[FUEL_SHEET])

            if FUEL_SHEET in wb_region.sheetnames:
                ws = wb_region[FUEL_SHEET]
                results["region_summary"] = analyze_region_fuel(ws)

    except Exception as e:
        results["errors"].append(f"Region report error: {str(e)}")

    return results