import openpyxl

MONTHLY_DATA_START_ROW = 8
MONTHLY_TREND_WINDOW = 4
MONTHLY_ACCOUNT_MIN_GAL = 25000


def safe_float(val):
    if val is None or val == "" or val == "%" or val == "#":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def clean_field(val):
    if val is None or val == "#" or str(val).strip() == "#":
        return None
    return str(val).strip() if val else None


def pct_change(current, prior):
    if prior is None or prior == 0 or current is None:
        return None
    return ((current - prior) / prior) * 100


def get_period(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 2:
            for c in (1, 2):
                if c < len(row) and row[c]:
                    return str(row[c])
    return "Unknown"


def analyze_regional_rollup(ws):
    regions = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < MONTHLY_DATA_START_ROW:
            continue
        label = clean_field(row[0])
        rep = clean_field(row[1])
        dsl = safe_float(row[2])
        yoy_dsl = safe_float(row[10])
        tires = safe_float(row[6])
        pm = safe_float(row[8])
        labor = safe_float(row[9])
        profit = safe_float(row[17])
        ppg = safe_float(row[19])
        acpg = safe_float(row[21])

        if not label or label in ("Customer", "Fleet Hierarchy", ""):
            continue
        if dsl is None:
            continue

        yoy_pct = None
        if yoy_dsl is not None and dsl is not None:
            prior_year = dsl - yoy_dsl
            if prior_year and prior_year != 0:
                yoy_pct = (yoy_dsl / prior_year) * 100

        regions.append({
            "label": label,
            "rep": rep,
            "dsl": dsl,
            "yoy_dsl": yoy_dsl,
            "yoy_pct": yoy_pct,
            "tires": tires,
            "pm": pm,
            "labor": labor,
            "profit": profit,
            "ppg": ppg,
            "acpg": acpg,
        })
    return regions


def analyze_area_manager_rollup(ws):
    managers = []
    current_group = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < MONTHLY_DATA_START_ROW:
            continue
        group = clean_field(row[0])
        mgr = clean_field(row[1])
        dsl = safe_float(row[2])
        yoy_dsl = safe_float(row[10])

        if group and group not in ("Overall Result",):
            current_group = group
        if not mgr or dsl is None:
            continue

        yoy_pct = None
        if yoy_dsl is not None and dsl is not None:
            prior_year = dsl - yoy_dsl
            if prior_year and prior_year != 0:
                yoy_pct = (yoy_dsl / prior_year) * 100

        managers.append({
            "manager": mgr,
            "group": current_group,
            "dsl": dsl,
            "yoy_dsl": yoy_dsl,
            "yoy_pct": yoy_pct,
        })
    managers.sort(key=lambda x: x["dsl"], reverse=True)
    return managers


def analyze_trend(ws):
    declining = []
    month_cols = list(range(3, 3 + MONTHLY_TREND_WINDOW))
    month_labels = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 4:
            month_labels = [clean_field(row[c]) for c in month_cols]
            break

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 7:
            continue
        region = clean_field(row[0])
        customer = clean_field(row[1])
        if not customer or customer in ("Overall Result",):
            continue

        vals = [safe_float(row[c]) for c in month_cols]
        if any(v is None for v in vals):
            continue
        is_declining = all(vals[k] < vals[k+1] for k in range(len(vals)-1))
        if is_declining and vals[-1] and vals[-1] >= MONTHLY_ACCOUNT_MIN_GAL:
            total_drop = vals[-1] - vals[0]
            drop_pct = pct_change(vals[0], vals[-1])
            declining.append({
                "region": region,
                "customer": customer,
                "recent_month": vals[0],
                "window_start": vals[-1],
                "drop_vol": total_drop,
                "drop_pct": drop_pct,
                "months": vals,
            })
    declining.sort(key=lambda x: x["drop_vol"], reverse=True)
    return declining, month_labels


def analyze_monthly_report(path):
    results = {
        "period": None,
        "regional": [],
        "area_managers": [],
        "trend_declining": [],
        "trend_months": [],
        "trend_window": MONTHLY_TREND_WINDOW,
        "errors": [],
    }
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        if "Sales Summary - Fuel" in wb.sheetnames:
            ws = wb["Sales Summary - Fuel"]
            results["period"] = get_period(ws)
            results["regional"] = analyze_regional_rollup(ws)

        if "Sales Summary - Area Manager" in wb.sheetnames:
            results["area_managers"] = analyze_area_manager_rollup(wb["Sales Summary - Area Manager"])

        if "13 Months History" in wb.sheetnames:
            declining, months = analyze_trend(wb["13 Months History"])
            results["trend_declining"] = declining
            results["trend_months"] = months

    except Exception as e:
        results["errors"].append(f"Monthly report error: {str(e)}")

    return results